import pandas as pd
import xlwings as xw 
import openpyxl
from openpyxl import load_workbook
from warnings import filterwarnings
import os 
# from multiprocessing.dummy import Pool as ThreadPool
from multiprocessing import Pool as ThreadPool
import time
import streamlit as st


#此代码用于定制 适用2023试算模板

'''获取文件夹下所有xlsx文件 '''
def get_file_list(folder_path,mode):
    file_list = []
    if mode=='穿透文件夹':
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if (file.endswith('.xlsx') or file.endswith('.xls') or file.endswith('.xlsm')) and '试算' not in file and '~$' not in file and '合并' not in file and '日志' not in file:
                    file_list.append(os.path.join(root, file))
                else:
                    continue
    elif mode=='非穿透':
        path_list = os.listdir(folder_path)
        file_list=[os.path.join(folder_path,file) for file in path_list if (file.endswith('.xlsx') or file.endswith('.xls') or file.endswith('.xlsm')) and '~$' not in file and '合并' not in file and '日志' not in file and '统计' not in file]
    
    return file_list


def extract_sheet_to_df_dict(source_file):
    '''提取试算 sheet 到存放dataframe的字典'''
    filterwarnings('ignore')  # 忽略警告信息
    file_name = os.path.basename(source_file) #提取文件名称
    file_name = file_name.replace('.xlsx','') #去掉后缀名

    sheet_name_list = ['CF2', '应收账款', '应付账款', '预收账款', '预付账款', '其他非流动资产','合同负债']

    result_dict={}
    # 打开源Excel文件
    try:
        wb_source = openpyxl.load_workbook(source_file,data_only=True,read_only=True)
    except Exception as e:
        #返回错误信息
        return {key:pd.DataFrame() for key in sheet_name_list}

    for sheet_name in sheet_name_list:
        try:
            # 获取指定的Sheet
            sheet = wb_source[sheet_name]
            # 复制数据
            data=[]
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            df=pd.DataFrame(data)
            raw_col=[i for i in range(0,len(df.columns))]
            df['文件名']=file_name
            new_col=['文件名']+raw_col # 文件名 放第一列
            df = df[new_col]
            if sheet_name in ['CF2']:
                #若模板不带外币
                if len(df.columns)<42: 
                    #1.补充字段
                    for num in range(35,43):
                        df[num]=''
                    # 2.把24-34 挪到28-38
                    for num_2 in range(28,39):
                        df[num_2]=df[num_2-4]
                    # 3.原币 币种 汇率 校验  补充
                    df[24]=''
                    df[25]='CNY'
                    df[26]=1
                    df[27]=''
                    df[39]=''
                    df[40]='CNY'
                    df[41]=''
                    df[42]=''
                else:
                    pass
            elif sheet_name in ['应付账款','预付账款','其他非流动资产']:
                if len(df.columns)<16:
                    #1.补充字段
                    for num in range(12,16):
                        df[num]=''
                    # 2.原币 币种 汇率 校验  补充
                    df[11]=''
                    df[12]='CNY'
                    df[13]=1
                    df[14]=''
            elif sheet_name in ['应收账款']:
                if len(df.columns)<19:
                    #1.补充字段
                    for num in range(14,19):
                        df[num]=''
                    #  2.是否单项认定计提坏账准备 	 原币 	 币种 	 汇率 	 校验 补充
                    df[14]=''
                    df[15]=''
                    df[16]='CNY'
                    df[17]=1
                    df[18]=''
            else:
                pass

            result_dict[sheet_name]=df # 存入字典 sheet_name 为往来科目名称 

        except Exception as e:
            result_dict[sheet_name]=pd.DataFrame()


    
    return result_dict



def clean_df_cf2(df):
    '''清洗CF2数据'''
    #拆分其他应收其他应付
    #其他应收 
    col_oar=['文件名']+[i for i in range(11,28)] #这里改成对应的列
    df_oar=df.iloc[4:,:][col_oar] #前4行不读
    df_oar=df_oar[~df_oar[11].isin(['其他应收款','名称'])] #删除辅助行
    df_oar=df_oar[df_oar[11].notnull()]#删除辅助行
    df_oar=df_oar[(df_oar[13].notnull())|(df_oar[14].notnull())] # 期初余额or期末余额 不同时为空
    #重命名表头
    df_oar.columns=['文件名',
                     '名称',
                    '款项性质',
                    '期初余额',
                    '期末余额',
                    '1年以内',
                    '1-2年',
                    '2-3年',
                    '3-4年',
                    '4-5年',
                    '5年以上',
                    '校验_1',
                    '收到往来款',
                    '支付往来款',
                    '原币',
                    '币种',
                    '汇率',
                    '校验_2']

    #其他应付 
    col_oap=['文件名']+[i for i in range(28,43)]
    df_oap=df.iloc[4:,:][col_oap] #前4行不读
    df_oap=df_oap[~df_oap[28].isin(['其他应付款','名称'])]
    df_oap=df_oap[df_oap[28].notnull()]
    df_oap=df_oap[(df_oap[30].notnull())|(df_oap[31].notnull())] # 期初余额or期末余额 不为空
    #重命名表头
    df_oap.columns=['文件名',
                     '名称',
                    '款项性质',
                    '期初余额',
                    '期末余额',
                    '1年以内',
                    '1-2年',
                    '2-3年',
                    '3年以上',
                    '校验_1',
                    '收到往来款',
                    '支付往来款',
                    '原币',
                    '币种',
                    '汇率',
                    '校验_2']

    return df_oar,df_oap



def clean_df(df,sheet_name):
    '''分情况清洗往来数据 folder_name就是sheet_name '''
    if sheet_name in ['应付账款','预付账款','其他非流动资产']:
        #0.选择需要的列
        temp_col=['文件名']+[i for i in range(0,16)] #这里改成对应的列
        df=df.iloc[5:,:][temp_col] #前5行不读
        #1.删除辅助行
        df=df[~df[0].isin(['返回科目披露','应付账款余额表','数据提供人','预付账款余额表','其他非流动资产余额表','合计'])]
        df=df[df[6].notnull()&(df[6]!=0)]#期末余额不为0 不为空
        #2.重命名表头
        df.columns=['文件名',
                    '数据提供人',
                    '被审计单位简称',
                    '客户/供应商名称',
                    '合并关联方',
                    '非合并关联方',
                    '款项性质',
                    '期末余额',
                    '1年以内',
                    '1-2年',
                    '2-3年',
                    '3年以上',
                    '合计',
                    '原币',
                    '币种', 
                    '汇率', 
                    '校验']
    elif sheet_name in ['应收账款']:
        #0.选择需要的列
        temp_col=['文件名']+[i for i in range(0,19)] #这里改成对应的列
        df=df.iloc[5:,:][temp_col] #前5行不读
        #1.删除辅助行
        df=df[~df[0].isin(['返回科目披露','应收账款余额表','助记码','合计'])]
        df=df[df[6].notnull()&(df[6]!=0)]#期末余额不为0 不为空
        #2.重命名表头
        df.columns=['文件名',
                    '助记码',
                    '被审计单位简称',
                    '客户/供应商名称',
                    '合并关联方',
                    '非合并关联方',
                    '款项性质',
                    '期末余额',
                    '1年以内',
                    '1-2年',
                    '2-3年',
                    '3-4年',
                    '4-5年',
                    '5年以上',
                    '合计',
                    '是否单项认定计提坏账准备',
                    '原币',
                    '币种', 
                    '汇率', 
                    '校验']
    elif sheet_name in ['合同负债']:
        #0.选择需要的列
        temp_col=['文件名']+[i for i in range(0,16)] #这里改成对应的列
        df=df.iloc[5:,:][temp_col] #前5行不读
        #1.删除辅助行
        df=df[~df[0].isin(['返回科目披露','应收账款余额表','数据提供人','助记码','合计'])]
        df=df[df[6].notnull()&(df[6]!=0)]#期末余额不为0 不为空
        #2.重命名表头
        df.columns=['文件名',
                    '助记码',
                    '被审计单位简称',
                    '客户/供应商名称',
                    '合并关联方',
                    '非合并关联方',
                    '款项性质',
                    '期末余额',
                    '1年以内',
                    '1-2年',
                    '2-3年',
                    '3年以上',
                    '合计',
                    '外销/内销',
                    '税率',
                    '合同负债',  
                    '其他流动负债']
    else:
        pass

    # elif sheet_name in ['预收账款']:
    #     #0.只要前18列
    #     df=df.iloc[3:,:18]
    #     #1.删除辅助行
    #     df=df[~df[0].isin(['客户编码','合计','下方填写合并关联方','下方填写非合并关联方','下方填写其他'])]
    #     df=df[~df[5].isin(['币种'])]
    #     df=df[df[6].notnull()&(df[6]!=0)]#人民币金额不为0 不为空
    #     #2.重命名表头
    #     df.columns=['文件名',
    #                 '客户编码',
    #                 '客户名称',
    #                 '合并关联方（是/不填）',
    #                 '非合并关联方（是/不填）',
    #                 '期初余额_原币',
    #                 '币种',
    #                 '人民币金额',
    #                 '1年以内',
    #                 '1-2年',                    
    #                 '2-3年',
    #                 '3年以上',
    #                 '余额合计核对',
    #                 '外币金额核对',
    #                 '余额由外币和人民币组成对，其中人民币金额',
    #                 '税率',
    #                 '合同负债',
    #                 '其他流动负债']

    # elif sheet_name in ['其他非流动资产']:
    #     #0.只要前18列
    #     df=df.iloc[3:,:18]
    #     #1.删除辅助行
    #     df=df[~df[0].isin(['合计','下方填写合并关联方','下方填写非合并关联方','下方填写其他'])]
    #     df=df[~df[6].isin(['币种'])]
    #     df=df[df[7].notnull()&(df[7]!=0)]#人民币金额不为0 不为空
    #     #2.重命名表头
    #     df.columns=['文件名',
    #                 '供应商编码',
    #                 '供应商名称',
    #                 '合并关联方（是/不填）',
    #                 '非合并关联方（是/不填）',
    #                 '款项性质',
    #                 '期初余额_原币'
    #                 '币种',
    #                 '人民币金额',
    #                 '1年以内',
    #                 '1-2年',                    
    #                 '2-3年',
    #                 '3年以上',
    #                 '余额合计核对',
    #                 '外币金额核对',
    #                 '余额由外币和人民币组成对，其中人民币金额']
    
    return df

# 批量从原始Excel文件提取往来并合并到新文件
def main_merge_raw_wb(source_path,save_folder,mode):
    
    filterwarnings('ignore')  # 忽略警告信息
    source_path=source_path

    #判断保存文件夹是否存在，不存在则创建
    os.makedirs(save_folder,exist_ok=True)

    #1.获取excel文件列表
    file_list = get_file_list(source_path,mode) # 默认不读打开的和有"合并"字样 的文件
    args=zip(file_list)

    #2.批量提取文件的各往来科目{'sheet_name':df,}到列表   
    print("开始提取试算底稿数据,请耐心等待...")
    cpu_count = os.cpu_count()
    with ThreadPool(processes=cpu_count) as pool:
        dict_list=pool.starmap(extract_sheet_to_df_dict, args)
    
    #3.合并相同key的dataframe
    sheet_name_list = ['CF2', '应收账款', '应付账款', '预收账款', '预付账款', '其他非流动资产','合同负债']
    for sheet_name in sheet_name_list:
        df_list=[temp_dict[sheet_name] for temp_dict in dict_list]
        df=pd.concat(df_list)
        #分情况清洗，如果是CF2要拆出来其他应收和其他应付
        if sheet_name == 'CF2':
            df_oar,df_oap=clean_df_cf2(df)  
            df_oar.to_excel(os.path.join(save_folder,"其他应收款.xlsx"),index=False)
            df_oap.to_excel(os.path.join(save_folder,"其他应付款.xlsx"),index=False)
        else:#按格子表格情况处理
            # xw.view(df)
            df=clean_df(df,sheet_name)
            df.to_excel(os.path.join(save_folder,f"{sheet_name}.xlsx"),index=False)
        #进度条
        print(f'进度:{(sheet_name_list.index(sheet_name)+1)/len(sheet_name_list)*100}%')
        print(f"合并【{sheet_name}】完成")
    print(f"全部合并完成,详见{save_folder}文件夹")



if __name__ == '__main__':

    # source_path=r'C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \华峰化学2024年报\2、试算、报告、小结\1、试算表\2、年审'
    # save_folder=r"D:\audit_project\试算提取\华峰"
    # main_merge_raw_wb(source_path,save_folder)

    # source_path=r'C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \东方生物2024年年审\2、试算\2024年试算-最新'
    # save_folder=r"D:\audit_project\AUTO_TB\dist"
    # main_merge_raw_wb(source_path,save_folder,"穿透文件夹")
    # print(get_file_list(source_path,"穿透文件夹"))


    source_path=r"C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \戈尔德2024年报\1、试算\2023年试算"
    save_folder=r"C:\Users\yefan\WPSDrive\339514258\WPS云盘\戈尔德\2023往来"
    main_merge_raw_wb(source_path,save_folder,"非穿透")
