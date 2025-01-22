import pandas as pd 
import xlwings as xw 
import duckdb
from openpyxl import load_workbook
from warnings import filterwarnings
import time
from win32com.client import Dispatch 
from multiprocessing import Pool
import re
from multiprocessing.dummy import Pool as ThreadPool
from concurrent.futures import ProcessPoolExecutor
import os,sys
sys.path.append(os.getcwd())
from module.tool_fun import get_file_list



# 参考 试算底稿路径
# r"C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \东方生物2024年年审\2、试算\2024年试算-最新\1.00 浙江东方基因生物制品股份有限公司 2024.xlsx"

'''
功能描述：
1.从试算取数
2.计算结果
3.把对应数据粘贴到底稿模板位置

模板位置：
r"D:\audit_project\AUTO_TB\DATA\期间费用模板.xlsx"
'''

#####################################自定义从底稿取数，粘贴到excel指定位置()区域需要连续############################################

'''从excel文件取数 表头仅接受一行的情况 '''
def get_data_from_paper(path,sheet_name,start_cell,end_cell,engine,header=None):
    path = path
    sheet_name = sheet_name
    start_cell = start_cell
    end_cell = end_cell
    engine = engine
    header = 0 if header is None else header

    company_name = os.path.basename(path).replace('.xlsx','')

    filterwarnings('ignore')
    df=pd.DataFrame()

    #根据不同的引擎读取数据
    try:
        if engine == 'openpyxl':
            wb = load_workbook(path,data_only=True,read_only=True) #只读数据不读公式
            sheet = wb[sheet_name]
            data = []
            for row in sheet[f'{start_cell}:{end_cell}']:
                data.append([cell.value for cell in row])

            col_names = data[header]#首行默认为列名
            df = pd.DataFrame(data[header+1:],dtype=object,columns=col_names)  # 设置列名
            

        elif engine =='excel':
            with xw.App(visible=False,add_book=False) as app:
                wb = app.books.open(path)
                sheet = wb.sheets[sheet_name]
                data = sheet.range(f'{start_cell}:{end_cell}').value
                col_names = data[header]
                df = pd.DataFrame(data[header+1:],dtype=object,columns=col_names)  # 设置列名
                wb.close()

        elif engine=='wps':
            xl = xw._xlwindows.COMRetryObjectWrapper(Dispatch("Ket.Application")) 
            impl = xw._xlwindows.App(visible=False, add_book=False, xl=xl)     
            with xw.App(visible=False, add_book=False, impl=impl) as app:
                wb = app.books.open(path)
                sheet = wb.sheets[sheet_name]
                data = sheet.range(f'{start_cell}:{end_cell}').value
                col_names = data[header]
                df = pd.DataFrame(data[header+1:],dtype=object,columns=col_names)  # 设置列名
                wb.close()

    except Exception as e:
        print(f'从{path}的{sheet_name}的{start_cell}到{end_cell}取数失败，错误信息：{e}')

    #把公司名放到前面那列
    raw_col=df.columns.tolist()
    df['公司']=company_name
    colnames=['公司']+raw_col
    df=df[colnames].copy()

    return df



'''粘贴数据到底稿'''
def paste_workingpaper(df,path_paper,sheet_name,first_cell,engine=None,path_save=None):

    path_paper = path_paper # 底稿路径
    sheet_name = sheet_name # 底稿sheet名
    first_cell = first_cell # 数据粘贴的第一个单元格
    start_row = int(re.findall('\d+',first_cell)[0])+1 #从单元格下一行加行
    
    df = df.copy() # 需要粘贴的数据

    try:
        if engine =='excel' or engine is None: #默认使用excel引擎
            with xw.App(visible=False,add_book=False) as app:
                wb = app.books.open(path_paper)
                sheet = wb.sheets[sheet_name]

                row_count = len(df) # DataFrame 的行数 要插入的行数
                # 复制并插入多行
                sheet.api.Rows(f"{start_row+1}:{start_row+1}").Copy()
                sheet.api.Rows(f"{start_row}:{start_row + row_count - 1}").Insert(
                    Shift=-4121  # 等效于 xlDown
                )
                sheet.range(f'{first_cell}').value = df.values.tolist()
                if path_save is not None:
                    wb.save(path_save)
                else:
                    wb.save()
                wb.close()
        else:
            xl = xw._xlwindows.COMRetryObjectWrapper(Dispatch("Ket.Application")) 
            impl = xw._xlwindows.App(visible=False, add_book=False, xl=xl)     
            with xw.App(visible=False, add_book=False, impl=impl) as app:
                wb = app.books.open(path_paper)
                sheet = wb.sheets[sheet_name]

                # 定义插入行的起始位置
                row_count = df.shape[0]  # DataFrame 的行数
                # 复制并插入多行
                sheet.api.Rows(f"{start_row+1}:{start_row+1}").Copy()
                sheet.api.Rows(f"{start_row}:{start_row + row_count - 1}").Insert(
                    Shift=-4121  # 等效于 xlDown
                )
                sheet.range(f'{first_cell}').value = df.values.tolist()
                if path_save is not None:
                    wb.save(path_save)
                else:
                    wb.save()
                wb.close()
    except Exception as e:
        print(f'粘贴数据到{path_paper}的{sheet_name}的{start_row}行{first_cell}单元格失败，错误信息：{e}')



'''
从excel批量取数 多进程
默认读取有'公司';非日志;非打开的excel文件
'''
def read_excel_multi(path,sheet_name,start_cell,end_cell,engine,header):
    #进程池 
    cpu_count = os.cpu_count()
    
    temp_path_list = get_file_list(path)
    # 默认读取有'公司';非日志;非打开的excel文件
    path_list = [i for i in temp_path_list if ('公司' in i and '日志' not in i and '~$' not in i and '小合并' not in i)]
    sheet_name_list=[sheet_name for i in range(len(path_list))]
    start_cell_list=[start_cell for i in range(len(path_list))]
    end_cell_list=[end_cell for i in range(len(path_list))]
    engine_list=[engine for i in range(len(path_list))]
    header_list=[header for i in range(len(path_list))]
    args=zip(path_list,sheet_name_list,start_cell_list,end_cell_list,engine_list,header_list)
    # with Pool(processes=cpu_count) as pool:
    #     result=pool.starmap(get_data_from_paper,args)
    # with ProcessPoolExecutor(max_workers=cpu_count) as executor:
    #     result=list(executor.starmap(get_data_from_paper,args))
    with ThreadPool(processes=cpu_count) as pool:
        result=pool.starmap(get_data_from_paper,args)
    df=pd.concat(result)

    return df

'''自定义取数粘贴'''
def custom_read_and_paste_main(path_from,sheet_name_from,start_cell_from,end_cell_from,
                               path_to,sheet_name_to,start_cell_to,engine,path_save):                        
    para_get_data={
        'path': path_from,
        'sheet_name': sheet_name_from,
        'start_cell': start_cell_from,
        'end_cell': end_cell_from,
        'engine':'openpyxl',
        'header':None
    }
    #1.批量读取数据
    df=read_excel_multi(**para_get_data)

    #2.粘贴数据到底稿指定位置
    para_paste_data={
        'df':df,
        'path_paper':path_to,
       'sheet_name':sheet_name_to,
        'first_cell':start_cell_to,
        'engine':engine,
        'path_save':path_save
    }
    paste_workingpaper(**para_paste_data)



##############################################自动生成费用底稿##################################################
'''
取各种费用（销售费用、管理费用、研发费用）数据
返回一个字典
'''
def get_cost_data(path):
    path=path
    filterwarnings('ignore')
    result={}
    
    #使用openpyxl读取数据 区域需要包含表头
    def get_data_openpyxl(wb,sheet_name,start_cell,end_cell,header=None):
        sheet = wb[sheet_name]
        data = []
        header = 0 if header is None else header #表头默认首行
        for row in sheet[f'{start_cell}:{end_cell}']:
            data.append([cell.value for cell in row])
        col_names = data[header]#首行默认为列名
        df = pd.DataFrame(data[header+1:],dtype=object,columns=col_names)  # 设置列名
        
        #!!!!把"公司"提到列名前面!!!!
        raw_colnames=df.columns.tolist()
        df['公司']=path.split('\\')[-1].replace('.xlsx','')
        col_names=['公司']+raw_colnames
        df=df[col_names]

        return df
    #审定表取数 ['公司','科目编码','底稿科目','附注分类','本期未审','重分类调整','审计调整']
    #销售费用 B1:G61
    #管理费用 B67:G131
    #研发费用 B137:G201
    wb = load_workbook(path,data_only=True,read_only=True) #只读数据不读公式

    df_SA_expenses=get_data_openpyxl(wb, '8_费用', 'B1', 'G61', header=0)
    df_GA_expenses=get_data_openpyxl(wb, '8_费用', 'B67', 'G131', header=0)
    df_RD_expenses=get_data_openpyxl(wb, '8_费用', 'B137', 'G201', header=0)

    #同期比较取数 ['公司','项目','本期发生额','上期发生额']
    #销售费用 A93:C123 
    #管理费用 A126:C156 
    #研发费用 A159:C189
    df_SA_com=get_data_openpyxl(wb, '8', 'A93', 'C123', header=0)
    df_GA_com=get_data_openpyxl(wb, '8', 'A126', 'C156', header=0)
    df_RD_com=get_data_openpyxl(wb, '8', 'A159', 'C189', header=0)

    result['审定表_管理费用']=df_GA_expenses
    result['审定表_研发费用']=df_RD_expenses
    result['审定表_销售费用']=df_SA_expenses

    result['同期比较_管理费用']=df_GA_com
    result['同期比较_研发费用']=df_RD_com
    result['同期比较_销售费用']=df_SA_com

    return result

'''1.批量读取数据'''
def get_cost_data_multi(path):

    cpu_count = os.cpu_count()
    path=path
    
    temp_path_list = get_file_list(path)
    path_list = [i for i in temp_path_list if ('公司' in i and '日志' not in i and '~$' not in i and '小合并' not in i)] #默认读取有'公司';非日志;非打开的excel文件
    args=path_list
    # 进程池
    # with Pool(processes=cpu_count-2) as pool:
        # results=executor.map(get_cost_data,args) #返回各公司字典[result1,result2,result3]

    # with ProcessPoolExecutor(max_workers=cpu_count) as executor:
    #     results=list(executor.map(get_cost_data,args)) #返回各公司字典[result1,result2,result3]

    with ThreadPool(processes=cpu_count) as pool:
        results=pool.map(get_cost_data,args) #返回各公司字典[result1,result2,result3]

    return results

'''2.粘贴数据到底稿'''
def paste_data(results,path,path_save):

    #粘贴审定表
    def paste_df_to_excel(df,wb,sheet_name,first_cell,title,new_sheet_name,copy,add_col):
        df = df.copy() # 需要粘贴的数据
        wb = wb # 底稿excel文件对象
        sheet_name = sheet_name # 底稿sheet名
        first_cell = first_cell # 数据粘贴的第一个单元格
        start_row = int(re.findall('\d+',first_cell)[0])+1 #从单元格下一行加行
        new_sheet_name = sheet_name if new_sheet_name is None else new_sheet_name
        
        try:
            #复制原sheet并重命名
            if copy==True:
                raw_sheet = wb.sheets[sheet_name]
                sheet = raw_sheet.copy(after=raw_sheet)
                sheet.name=new_sheet_name
            else:
                sheet = wb.sheets[sheet_name]

            #加标题
            if title is not None:
                sheet.range('A1').value = title

            # 加行
            row_count = len(df) # DataFrame 的行数 要插入的行数
            if add_col==True:
                # 复制并插入多行
                sheet.api.Rows(f"{start_row+1}:{start_row+1}").Copy()
                sheet.api.Rows(f"{start_row}:{start_row + row_count - 1}").Insert(
                    Shift=-4121  # 等效于 xlDown
                )
            # 粘贴数据
            sheet.range(f'{first_cell}').value = df.values.tolist()
        except Exception as e:
            print(f'粘贴数据到{wb.name}的{sheet_name}的{start_row}行{first_cell}单元格失败，错误信息：{e}')

    #审定表取数
    SA_list=[i['审定表_销售费用'] for i in results]
    GA_list=[i['审定表_管理费用'] for i in results]
    RD_list=[i['审定表_研发费用'] for i in results]

    #同期比较取数
    SA_com_list=[i['同期比较_销售费用'] for i in results]
    GA_com_list=[i['同期比较_管理费用'] for i in results]
    RD_com_list=[i['同期比较_研发费用'] for i in results]

    #合并数据 ['公司','科目编码','底稿科目','附注分类','本期未审','重分类调整','审计调整']
    SA_df=pd.concat(SA_list)
    GA_df=pd.concat(GA_list)
    RD_df=pd.concat(RD_list)

    #合并数据 ['公司','项目','本期发生额','上期发生额']
    SA_com_df=pd.concat(SA_com_list)
    GA_com_df=pd.concat(GA_com_list)
    RD_com_df=pd.concat(RD_com_list)

    #清洗数据
    SA_df=SA_df[SA_df['本期未审'].notnull()].copy()
    GA_df=GA_df[GA_df['本期未审'].notnull()].copy()
    RD_df=RD_df[RD_df['本期未审'].notnull()].copy()

    # wps打开excel文件，粘贴数据    
    # xl = xw._xlwindows.COMRetryObjectWrapper(Dispatch("Ket.Application")) 
    # impl = xw._xlwindows.App(visible=False, add_book=False, xl=xl)     
    # with xw.App(visible=False, add_book=False, impl=impl) as app:

    # excel打开 excel保存 wps没法保存
    with xw.App(visible=False, add_book=False) as app:
        wb = app.books.open(path)
        #贴审定表
        paste_df_to_excel(df=SA_df,wb=wb,sheet_name='审定表',first_cell='A9',title='销售费用审定表',new_sheet_name='审定表_销售费用',copy=True,add_col=True)
        paste_df_to_excel(df=GA_df,wb=wb,sheet_name='审定表',first_cell='A9',title='管理费用审定表',new_sheet_name='审定表_管理费用',copy=True,add_col=True)
        paste_df_to_excel(df=RD_df,wb=wb,sheet_name='审定表',first_cell='A9',title='研发费用审定表',new_sheet_name='审定表_研发费用',copy=True,add_col=True)
        #贴同期比较
        paste_df_to_excel(df=SA_com_df[['公司','项目','本期发生额']],wb=wb,sheet_name='同期比较',first_cell='A9',title='销售费用本期与上年同期比较表',new_sheet_name='同期比较_销售费用',copy=True,add_col=True)
        paste_df_to_excel(df=GA_com_df[['公司','项目','本期发生额']],wb=wb,sheet_name='同期比较',first_cell='A9',title='管理费用本期与上年同期比较表',new_sheet_name='同期比较_管理费用',copy=True,add_col=True)
        paste_df_to_excel(df=RD_com_df[['公司','项目','本期发生额']],wb=wb,sheet_name='同期比较',first_cell='A9',title='研发费用本期与上年同期比较表',new_sheet_name='同期比较_研发费用',copy=True,add_col=True)

        #不加行直接贴
        paste_df_to_excel(df=SA_com_df[['上期发生额']],wb=wb,sheet_name='同期比较_销售费用',first_cell='E9',title=None,new_sheet_name=None,copy=False,add_col=False)
        paste_df_to_excel(df=GA_com_df[['上期发生额']],wb=wb,sheet_name='同期比较_管理费用',first_cell='E9',title=None,new_sheet_name=None,copy=False,add_col=False)
        paste_df_to_excel(df=RD_com_df[['上期发生额']],wb=wb,sheet_name='同期比较_研发费用',first_cell='E9',title=None,new_sheet_name=None,copy=False,add_col=False)

        #删除【省定表】【同期比较】
        wb.sheets['审定表'].delete()
        wb.sheets['同期比较'].delete()

        #保存
        if path_save is not None:
            wb.save(path_save)
        else:
            wb.save()
        wb.close()

'''生成费用底稿'''
def gen_cost_workingpaper(path_data,path_paper,path_save):

    path_data=path_data
    path_paper=path_paper
    path_save=path_save

    #1.批量读取数据
    results=get_cost_data_multi(path_data)

    #2.粘贴数据到底稿
    paste_data(results,path_paper,path_save)

        


if __name__ == '__main__':

    '''费用底稿贴数'''
    #参数设置
    start_time = time.time()
    path_data = r"C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \东方生物2024年年审\2、试算\2024年试算-最新"
    path_paper = r"D:\audit_project\AUTO_TB\示例数据\DATA\期间费用模板_empty.xlsx"
    path_save = r"D:\audit_project\AUTO_TB\示例数据\DATA\期间费用_FY24_东方生物.xlsx"
    gen_cost_workingpaper(path_data,path_paper,path_save)
    end_time = time.time()
    print(f'耗时：{round(end_time-start_time,2)}秒')
    print('done')


    '''自定义贴数'''
    #参数设置
    # path_from = r"C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \东方生物2024年年审\2、试算\2024年试算-最新"
    # sheet_name_from = '8_费用'
    # start_cell_from = 'B1'
    # end_cell_from = 'G61'

    # path_to = r"D:\audit_project\AUTO_TB\DATA\期间费用模板_empty.xlsx"
    # sheet_name_to = '审定表'
    # start_cell_to = 'A9'
    # engine = 'excel'
    # path_save=r"D:\audit_project\AUTO_TB\DATA\期间费用模板_test.xlsx"

    # custom_read_and_paste_main(path_from,sheet_name_from,start_cell_from,end_cell_from,
    #                            path_to,sheet_name_to,start_cell_to,engine,path_save)

    # print('done')


    




    #######################################草稿#####################################

    # freeze_support()
    # args={
    #     'path':r"C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \东方生物2024年年审\2、试算\2024年试算-最新",
    #     'sheet_name':'8_费用',
    #     'start_cell':'B1',
    #      'end_cell':'G61',
    #      'header':0,
    # }
    # result=read_excel_multi(**args)

    # result=result[['file_path','科目编码','底稿科目','附注分类','本期未审','重分类调整','审计调整']]
    # result.rename(columns={'file_path':'公司'},inplace=True)
    # result['公司']=result['公司'].apply(lambda x:os.path.basename(x).replace('.xlsx',''))
    # result=result[result['本期未审'].notnull()]

    # args={'df':result,
    #       'path_paper':r"D:\audit_project\AUTO_TB\DATA\期间费用模板_empty.xlsx",
    #       'sheet_name':'审定表 (2)',
    #       'start_row':10,
    #       'first_cell':'A9'}
    
    # paste_workingpaper(**args)

    
    # print('done')


    ################################################################



    # print(os.listdir(r'C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \东方生物2024年年审\2、试算\2024年试算-最新'))

    #销售费用 B1:G61
    #管理费用 B67:G131
    #研发费用 B137:G201
    # import os

    # print(os.path.basename(r'C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \东方生物2024年年审\2、试算\2024年试算-最新\2024年试算-销售费用.xlsx'))

    # path = r"C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \东方生物2024年年审\2、试算\2024年试算-最新\1.00 浙江东方基因生物制品股份有限公司 2024.xlsx"
    # para={
    # 'path':path,
    # 'sheet_name': '8_费用',
    # 'start_cell':'B67',
    # 'end_cell':'G113',
    # # 'show_time':True,
    # 'show_time':False,
    # # 'engine':'openpyxl',
    # # 'engine':'excel',
    # 'engine':'wps',
    # 'header':0
    # }
    # df = get_data_from_paper(**para)
    # xw.view(df)

    # print('hello world')






