import pandas as pd
import xlwings as xw 
import openpyxl
from openpyxl import load_workbook
from warnings import filterwarnings
import os 
from multiprocessing.dummy import Pool as ThreadPool
import time
import streamlit as st


'''获取文件夹下所有xlsx文件 '''
def get_file_list(folder_path,mode):
    file_list = []
    if mode=='穿透文件夹':
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if (file.endswith('.xlsx') or file.endswith('.xls') or file.endswith('.xlsm')) and '试算' not in file and '~$' not in file and '合并' not in file and '日志' not in file:
                    file_list.append(os.path.join(root, file))
                else:
                    continue
    elif mode=='非穿透':
        path_list = os.listdir(folder_path)
        file_list=[os.path.join(folder_path,file) for file in path_list if (file.endswith('.xlsx') or file.endswith('.xls') or file.endswith('.xlsm')) and '~$' not in file and '合并' not in file and '日志' not in file]
    
    return file_list


def extract_sheet_to_df_dict(source_file):
    '''提取试算 sheet 到存放dataframe的字典'''
    filterwarnings('ignore')  # 忽略警告信息
    file_name = os.path.basename(source_file) #提取文件名称
    file_name = file_name.replace('.xlsx','') #去掉后缀名

    sheet_name_list = ['CF2', '应收账款', '应付账款', '预收账款', '预付账款', '其他非流动资产']

    result_dict={}
    # 打开源Excel文件
    try:
        wb_source = openpyxl.load_workbook(source_file,data_only=True,read_only=True)
    except Exception as e:
        #返回错误信息
        return {key:pd.DataFrame() for key in sheet_name_list}
        # print(f'打开源文件{source_file}失败，错误信息：{e}')
    

    for sheet_name in sheet_name_list:
        try:
            # 获取指定的Sheet
            sheet = wb_source[sheet_name]
            # 复制数据
            data=[]
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            df=pd.DataFrame(data)
            raw_col=[i for i in range(0,len(df.columns))]
            df['文件名']=file_name
            new_col=['文件名']+raw_col # 文件名 放第一列
            df = df[new_col]
            result_dict[sheet_name]=df # 存入字典 sheet_name 为往来科目名称 

        except Exception as e:
            result_dict[sheet_name]=pd.DataFrame()


    
    return result_dict



def clean_df_cf2(df):
    '''清洗CF2数据'''
    #拆分其他应收其他应付
    #其他应收 前3行不读
    col_oar=['文件名']+[i for i in range(11,30)]
    df_oar=df.iloc[3:,:][col_oar]
    df_oar=df_oar[~df_oar[12].isin(['款项性质','合计','勾稽核对'])]
    df_oar=df_oar[df_oar[12].notnull()]#款项性质
    df_oar=df_oar[(df_oar[13].notnull())|(df_oar[14].notnull())|(df_oar[15].notnull())] # 期初余额or借方发生额or贷方发生额 不同时为空
    df_oar=df_oar[df_oar[16]!=0]
    #重命名表头
    df_oar.columns=['文件名',
                    '名称',
                    '款项性质',
                    '期初余额',
                    '借方发生额',
                    '贷方发生额',
                    '期末余额',
                    '单项坏账',
                    '收到往来款',
                    '支付往来款',
                    '1年以内',
                    '1-2年',
                    '2-3年',
                    '3-4年',
                    '4-5年',
                    '5年以上',
                    '外币金额',
                    '币种',
                    '账龄核对',
                    '外币核对']

    #其他应付 前3行不读
    col_oap=['文件名']+[i for i in range(30,46)]
    df_oap=df.iloc[3:,:][col_oap]
    df_oap=df_oap[~df_oap[31].isin(['款项性质','合计','勾稽核对'])]
    df_oap=df_oap[df_oap[31].notnull()]
    df_oap=df_oap[(df_oap[32].notnull())|(df_oap[33].notnull())|(df_oap[34].notnull())] # 期初余额or借方发生额or贷方发生额 不为空
    df_oap=df_oap[df_oap[35]!=0]
    #重命名表头
    df_oap.columns=['文件名',
                    '名称',
                    '款项性质',
                    '期初余额',
                    '借方发生额',
                    '贷方发生额',
                    '期末余额',
                    '收到往来款',
                    '支付往来款',
                    '1年以内',
                    '1-2年',
                    '2-3年',
                    '3年以上',
                    '外币金额',
                    '币种',
                    '账龄核对',
                    '外币核对']

    return df_oar,df_oap



def clean_df(df,sheet_name):
    '''分情况清洗往来数据 folder_name就是sheet_name '''
    if sheet_name in ['应付账款','预付账款','其他非流动资产']:
        #0.只要前16列
        df=df.iloc[4:,:16]
        #1.删除辅助行
        df=df[~df[0].isin(['供应商编码','合计','下方填写合并关联方','下方填写非合并关联方','下方填写其他'])]
        df=df[~df[6].isin(['币种'])]
        df=df[df[7].notnull()&(df[7]!=0)]#人民币金额不为0 不为空
        #2.重命名表头
        df.columns=['文件名',
                    '供应商编码',
                    '供应商名称',
                    '合并关联方（是/不填）',
                    '非合并关联方（是/不填）',
                    '款项性质',
                    '期末余额_原币',
                    '币种',
                    '人民币金额',
                    '1年以内',
                    '1-2年',
                    '2-3年',
                    '3年以上',
                    '余额合计核对',
                    '外币金额核对',
                    '余额由外币和人民币组成对，其中人民币金额'
                    ]
    elif sheet_name in ['应收账款']:
        #0.只要前18列
        df=df.iloc[3:,:18]
        #1.删除辅助行
        df=df[~df[0].isin(['客户编码','合计','下方填写合并关联方','下方填写非合并关联方','下方填写其他'])]
        df=df[~df[5].isin(['币种'])]
        df=df[df[6].notnull()&(df[6]!=0)]#人民币金额不为0 不为空
        #2.重命名表头
        df.columns=['文件名',
                    '客户编码',
                    '客户名称',
                    '合并关联方（是/不填）',
                    '非合并关联方（是/不填）',
                    '期初余额_原币',
                    '币种',
                    '人民币金额',
                    '1年以内',
                    '1-2年',                    
                    '2-3年',
                    '3-4年',
                    '4-5年',
                    '5年以上',
                    '是否单项认定计提坏账准备',
                    '余额合计核对',
                    '外币金额核对',
                    '余额由外币和人民币组成对，其中人民币金额'
                    ]
    elif sheet_name in ['预收账款']:
        #0.只要前18列
        df=df.iloc[3:,:18]
        #1.删除辅助行
        df=df[~df[0].isin(['客户编码','合计','下方填写合并关联方','下方填写非合并关联方','下方填写其他'])]
        df=df[~df[5].isin(['币种'])]
        df=df[df[6].notnull()&(df[6]!=0)]#人民币金额不为0 不为空
        #2.重命名表头
        df.columns=['文件名',
                    '客户编码',
                    '客户名称',
                    '合并关联方（是/不填）',
                    '非合并关联方（是/不填）',
                    '期初余额_原币',
                    '币种',
                    '人民币金额',
                    '1年以内',
                    '1-2年',                    
                    '2-3年',
                    '3年以上',
                    '余额合计核对',
                    '外币金额核对',
                    '余额由外币和人民币组成对，其中人民币金额',
                    '税率',
                    '合同负债',
                    '其他流动负债']

    elif sheet_name in ['其他非流动资产']:
        #0.只要前18列
        df=df.iloc[3:,:18]
        #1.删除辅助行
        df=df[~df[0].isin(['合计','下方填写合并关联方','下方填写非合并关联方','下方填写其他'])]
        df=df[~df[6].isin(['币种'])]
        df=df[df[7].notnull()&(df[7]!=0)]#人民币金额不为0 不为空
        #2.重命名表头
        df.columns=['文件名',
                    '供应商编码',
                    '供应商名称',
                    '合并关联方（是/不填）',
                    '非合并关联方（是/不填）',
                    '款项性质',
                    '期初余额_原币'
                    '币种',
                    '人民币金额',
                    '1年以内',
                    '1-2年',                    
                    '2-3年',
                    '3年以上',
                    '余额合计核对',
                    '外币金额核对',
                    '余额由外币和人民币组成对，其中人民币金额']
    
    return df

# 批量从原始Excel文件提取往来并合并到新文件
def main_merge_raw_wb(source_path,save_folder,mode):
    
    filterwarnings('ignore')  # 忽略警告信息
    source_path=source_path

    #判断保存文件夹是否存在，不存在则创建
    os.makedirs(save_folder,exist_ok=True)

    #1.获取excel文件列表
    file_list = get_file_list(source_path,mode) # 默认不读打开的和有"合并"字样 的文件
    args=zip(file_list)

    #2.批量提取文件的各往来科目{'sheet_name':df,}到列表   
    st.write("开始提取试算底稿数据,请耐心等待...")
    cpu_count = os.cpu_count()
    with ThreadPool(processes=cpu_count) as pool:
        dict_list=pool.starmap(extract_sheet_to_df_dict, args)
    
    #3.合并相同key的dataframe
    sheet_name_list = ['CF2', '应收账款', '应付账款', '预收账款', '预付账款', '其他非流动资产']
    for sheet_name in sheet_name_list:
        df_list=[temp_dict[sheet_name] for temp_dict in dict_list]
        df=pd.concat(df_list)
        #分情况清洗，如果是CF2要拆出来其他应收和其他应付
        if sheet_name == 'CF2':
            df_oar,df_oap=clean_df_cf2(df)  
            df_oar.to_excel(os.path.join(save_folder,"其他应收款.xlsx"),index=False)
            df_oap.to_excel(os.path.join(save_folder,"其他应付款.xlsx"),index=False)
        else:#按格子表格情况处理
            df=clean_df(df,sheet_name)
            df.to_excel(os.path.join(save_folder,f"{sheet_name}.xlsx"),index=False)
        #进度条
        st.progress((sheet_name_list.index(sheet_name)+1)/len(sheet_name_list))
        st.write(f"合并【{sheet_name}】完成")
    st.success(f"全部合并完成,详见文件夹: {save_folder}")



if __name__ == '__main__':

    # source_path=r'C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \华峰化学2024年报\2、试算、报告、小结\1、试算表\2、年审'
    # save_folder=r"D:\audit_project\试算提取\华峰"
    # main_merge_raw_wb(source_path,save_folder)

    source_path=r'C:\Users\yefan\WPSDrive\339514258\WPS云盘\共享文件夹 \东方生物2024年年审\2、试算\2024年试算-最新'
    save_folder=r"D:\audit_project\AUTO_TB\dist"
    main_merge_raw_wb(source_path,save_folder,"穿透文件夹")
    # print(get_file_list(source_path,"穿透文件夹"))
